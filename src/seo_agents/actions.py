from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from seo_agents.crew import OUTPUT_DIR


ACTION_QUEUE_FILE = OUTPUT_DIR / "action_queue.json"
ACTION_APPROVALS_FILE = OUTPUT_DIR / "action_approvals.json"
ACTION_COMPLETIONS_FILE = OUTPUT_DIR / "action_completions.json"
ACTION_COMPLETIONS_CONFIG_FILE = Path(os.getenv(
    "ACTION_COMPLETIONS_CONFIG_FILE",
    "config/action-completions.json",
))
ACTION_RUN_DIR = OUTPUT_DIR / "action_runs"
GBP_POSTER_SCRIPT = Path(os.getenv(
    "GBP_POSTER_SCRIPT",
    r"C:\Users\carte\.claude\skills\gbp-poster\driver.mjs",
))
GBP_POSTER_CONFIG = Path(os.getenv(
    "GBP_POSTER_CONFIG",
    r"C:\Users\carte\.codex\plugins\grizzly-gbp-poster\config.local.json",
))
GBP_BROWSER_SESSION_DIR = Path(os.getenv(
    "GBP_BROWSER_SESSION_DIR",
    r"C:\Users\carte\.claude\gbp-session",
))
WORDPRESS_SITE_CONFIG = Path(os.getenv(
    "WORDPRESS_SITE_CONFIG",
    r"C:\Workspace\Active\SEO-Agents-App\config\wordpress-sites\grizzly.json",
))
WORDPRESS_ACTION_ADAPTER = os.getenv(
    "WORDPRESS_ACTION_ADAPTER",
    r"C:\Workspace\Active\SEO-Agents-App\scripts\wordpress-action-adapter.mjs",
).strip()
WORDPRESS_BROWSER_SESSION_DIR = Path(os.getenv(
    "WORDPRESS_BROWSER_SESSION_DIR",
    r"C:\Workspace\Shared\Agents\BrowserSessions\grizzly-wordpress",
))
GBP_POSTER_TIMEOUT_S = int(os.getenv("GBP_POSTER_TIMEOUT_S", "420"))
GBP_POSTER_HEADLESS = os.getenv("GBP_POSTER_HEADLESS", "0").lower() in {"1", "true", "yes", "on"}
WORDPRESS_ADAPTER_TIMEOUT_S = int(os.getenv("WORDPRESS_ADAPTER_TIMEOUT_S", "300"))
GBP_PROFILE_ADAPTER = os.getenv(
    "GBP_PROFILE_ADAPTER",
    r"C:\Workspace\Active\SEO-Agents-App\scripts\gbp-profile-adapter.mjs",
).strip()
GBP_PROFILE_ADAPTER_TIMEOUT_S = int(os.getenv("GBP_PROFILE_ADAPTER_TIMEOUT_S", "300"))
FACEBOOK_POSTER_ADAPTER = os.getenv(
    "FACEBOOK_POSTER_ADAPTER",
    r"C:\Workspace\Active\SEO-Agents-App\scripts\facebook-poster.mjs",
).strip()
# Allow >720s so a slow Veo 3 video render (see gemini-video-generator MAX_POLL_ATTEMPTS)
# plus upload completes before this subprocess is killed.
FACEBOOK_POSTER_TIMEOUT_S = int(os.getenv("FACEBOOK_POSTER_TIMEOUT_S", "900"))
GBP_WORKBOOK_HEADERS = [
    "Date",
    "PostType",
    "Topic",
    "AssetSource",
    "AssetIdOrDescription",
    "CTA",
    "Status",
    "CaptionDraft",
    "ImageLink",
    "Posted",
    "PostedAt",
    "GBPPostUrl",
    "Notes",
]


def _now_iso() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def _markdown_body(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", stripped)
        stripped = re.sub(r"\s*```\s*$", "", stripped)
    return stripped.strip()


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(exist_ok=True)
    temp_path = path.with_suffix(f"{path.suffix}.tmp")
    temp_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    temp_path.replace(path)


def _run_subprocess(command: list[str], cwd: str, timeout_s: int, display_command: str | None = None) -> dict[str, Any]:
    """Run an adapter subprocess, returning a uniform result record.

    A hung browser run must come back as an adapter failure record (exit 124),
    not an uncaught TimeoutExpired blowing up run_action.
    """
    shown = display_command or " ".join(command)
    try:
        result = subprocess.run(
            command,
            cwd=cwd,
            capture_output=True,
            text=True,
            timeout=timeout_s,
            check=False,
        )
    except subprocess.TimeoutExpired as error:
        return {
            "exit_code": 124,
            "command": shown,
            "stdout": (error.stdout or b"").decode(errors="replace") if isinstance(error.stdout, bytes) else (error.stdout or ""),
            "stderr": f"Adapter timed out after {timeout_s}s.",
        }
    except OSError as error:
        return {
            "exit_code": 126,
            "command": shown,
            "stdout": "",
            "stderr": f"Failed to launch adapter: {error}",
        }
    return {
        "exit_code": result.returncode,
        "command": shown,
        "stdout": result.stdout.strip(),
        "stderr": result.stderr.strip(),
    }


def _last_json_line(stdout: str) -> dict[str, Any]:
    """Adapters emit a single-line JSON result as their final stdout line."""
    for line in reversed(stdout.splitlines()):
        line = line.strip()
        if line.startswith("{") and line.endswith("}"):
            try:
                return json.loads(line)
            except json.JSONDecodeError:
                continue
    return {}


def _extract_numbered_field(block: str, label: str) -> str:
    match = re.search(rf"\d+\)\s*\*\*{re.escape(label)}\*\*:\s*(.+)", block)
    return match.group(1).strip() if match else ""


def _extract_list_after(block: str, label: str) -> list[str]:
    marker = re.search(rf"\d+\)\s*\*\*{re.escape(label)}\*\*:\s*", block)
    if not marker:
        return []
    tail = block[marker.end():]
    next_field = re.search(r"^\d+\)\s*\*\*.+?\*\*:", tail, flags=re.MULTILINE)
    section = tail[:next_field.start()] if next_field else tail
    return [
        line.strip()[2:].strip()
        for line in section.splitlines()
        if line.strip().startswith("- ")
    ]


def _extract_completion_blocks(text: str) -> dict[str, dict[str, str]]:
    body = _markdown_body(text)
    parts = re.split(r"(?=^### COMPLETION REPORT)", body, flags=re.MULTILINE)
    completions: dict[str, dict[str, str]] = {}
    for part in parts:
        task_id = re.search(r"^Task ID:\s*([A-Z]+-?\d+)", part, flags=re.MULTILINE)
        if not task_id:
            continue
        completion: dict[str, str] = {"task_id": task_id.group(1)}
        for label, key in (
            ("Task Title", "title"),
            ("Assigned Agent", "agent"),
            ("Status", "completion_status"),
            ("Action Taken", "action_taken"),
            ("Deliverable Location", "deliverable_location"),
            ("Definition of Done Met", "definition_of_done"),
            ("If Partial or Blocked", "blocker"),
            ("Owner Sign-Off Needed", "owner_signoff_needed"),
        ):
            match = re.search(rf"^{re.escape(label)}:\s*(.+)", part, flags=re.MULTILINE)
            if match:
                completion[key] = match.group(1).strip()
        completions[completion["task_id"]] = completion
    task_parts = re.split(r"(?=^## Task \d+:)", body, flags=re.MULTILINE)
    for part in task_parts:
        task_id = re.search(r"^### Task ID:\s*([A-Z]+-?\d+)", part, flags=re.MULTILINE)
        if not task_id:
            continue
        completion = {"task_id": task_id.group(1)}
        title = re.search(r"^## Task \d+:\s*(.+)", part, flags=re.MULTILINE)
        if title:
            completion["title"] = title.group(1).strip()
        for label, key in (
            ("Status", "completion_status"),
            ("Definition of Done Met", "definition_of_done"),
            ("Deliverable Location", "deliverable_location"),
            ("If Partial or Blocked", "blocker"),
            ("Owner Sign-Off Needed", "owner_signoff_needed"),
        ):
            match = re.search(rf"^### {re.escape(label)}:\s*(.+)", part, flags=re.MULTILINE)
            if match:
                completion[key] = match.group(1).strip()
        action_match = re.search(
            r"^### Action Taken:\s*\n(.*?)(?=^### |\n---|\Z)",
            part,
            flags=re.MULTILINE | re.DOTALL,
        )
        if action_match:
            completion["action_taken"] = " ".join(action_match.group(1).split())
        completions[completion["task_id"]] = completion
    return completions


def _completions_from_json(payload: Any) -> dict[str, dict[str, str]]:
    """Parse a structured completion report (crew output_json schema)."""
    entries = payload.get("completions", []) if isinstance(payload, dict) else payload
    if not isinstance(entries, list):
        return {}
    completions: dict[str, dict[str, str]] = {}
    for entry in entries:
        if not isinstance(entry, dict) or not entry.get("task_id"):
            continue
        completions[str(entry["task_id"]).strip()] = {
            key: str(value).strip()
            for key, value in entry.items()
            if value is not None
        }
    return completions


def _load_completions() -> dict[str, dict[str, str]]:
    completions: dict[str, dict[str, str]] = {}
    for stem in ("content_completion", "assets_completion", "technical_completion"):
        # Structured JSON (preferred, regex-free) wins; markdown is the fallback
        # for older runs or models that failed structured output. A stale JSON
        # from a previous run must not shadow a fresher markdown report.
        json_path = OUTPUT_DIR / f"{stem}.json"
        md_path = OUTPUT_DIR / f"{stem}.md"
        if json_path.exists() and (not md_path.exists() or json_path.stat().st_mtime >= md_path.stat().st_mtime):
            parsed = _completions_from_json(_load_json(json_path, {}))
            if parsed:
                completions.update(parsed)
                continue
        completions.update(_extract_completion_blocks(_read_text(OUTPUT_DIR / f"{stem}.md")))
    return completions


def _load_action_completions() -> dict[str, dict[str, Any]]:
    """Load owner-verified / manual completion overrides.

    Keys are action IDs (e.g. ``task-t002``). Each entry must have a ``status``
    field (e.g. ``"verified"``) and optionally ``verified_at``, ``verified_by``,
    and ``details``.
    """
    completions = _load_json(ACTION_COMPLETIONS_CONFIG_FILE, {})
    completions.update(_load_json(ACTION_COMPLETIONS_FILE, {}))
    return completions


def _infer_action_type(executor: str, title: str, steps: list[str]) -> str:
    haystack = f"{executor} {title} {' '.join(steps)}".lower()
    if "contact form" in haystack or "technical" in executor.lower():
        return "website_technical_change"
    if "gbp" in haystack or "google business" in haystack:
        return "gbp_profile_update"
    if "facebook" in haystack or "fb" in executor.lower():
        return "publish_facebook_post"
    if "review" in haystack:
        return "review_management"
    if "blog post" in haystack or "blog" in executor.lower():
        return "website_blog_post"
    if "service page" in haystack or "content" in executor.lower():
        return "website_content_publish"
    return "manual_followup"


def _risk_for_action(action_type: str) -> str:
    if action_type in {"website_technical_change", "gbp_profile_update", "website_content_publish"}:
        return "high"
    if action_type in {"review_management", "publish_gbp_post", "website_blog_post", "publish_facebook_post"}:
        return "medium"
    return "low"


def _platform_for_action(action_type: str) -> str:
    return {
        "website_technical_change": "website_cms",
        "website_content_publish": "website_cms",
        "website_blog_post": "website_cms",
        "gbp_profile_update": "google_business_profile",
        "publish_gbp_post": "google_business_profile",
        "publish_facebook_post": "facebook_page",
        "review_management": "review_platforms",
    }.get(action_type, "manual")


def _status_for_action(
    completion: dict[str, str],
    dependencies: list[str],
    completion_override: dict[str, Any] | None = None,
) -> str:
    # Owner-verified / manual completion overrides short-circuit all other logic.
    if completion_override and completion_override.get("status") == "verified":
        return "verified"
    status = completion.get("completion_status", "").upper()
    dependency_text = " ".join(dependencies).lower()
    blocker_text = completion.get("blocker", "").lower()
    if status == "COMPLETE" and completion.get("definition_of_done", "").upper() == "YES":
        return "dry_run_ready"
    if "access" in dependency_text or "access" in blocker_text or status == "BLOCKED":
        return "blocked_access"
    if completion.get("owner_signoff_needed", "").upper() == "YES":
        return "needs_approval"
    if status in {"COMPLETE", "PARTIAL"}:
        return "dry_run_ready"
    return "needs_review"


def parse_execution_actions() -> list[dict[str, Any]]:
    queue_text = _markdown_body(_read_text(OUTPUT_DIR / "grizzly_execution_queue.md"))
    completions = _load_completions()
    completion_overrides = _load_action_completions()
    parts = re.split(r"(?=^## Task \d+:)", queue_text, flags=re.MULTILINE)
    actions: list[dict[str, Any]] = []
    for part in parts:
        task_id = _extract_numbered_field(part, "Task ID")
        if not task_id:
            continue
        title = _extract_numbered_field(part, "Task Title")
        executor = _extract_numbered_field(part, "Assigned Execution Agent")
        steps = _extract_list_after(part, "Exact Action Steps")
        dependencies = _extract_list_after(part, "Dependencies / Required Inputs")
        verification = _extract_list_after(part, "Verification Checklist")
        completion = completions.get(task_id, {})
        action_id = f"task-{task_id.lower()}"
        override = completion_overrides.get(action_id)
        action_type = _infer_action_type(executor, title, steps)
        platform = _platform_for_action(action_type)
        status = _status_for_action(completion, dependencies, override)
        actions.append({
            "id": action_id,
            "source": "execution_queue",
            "source_task_id": task_id,
            "title": title,
            "assigned_agent": executor,
            "action_type": action_type,
            "platform": platform,
            "risk": _risk_for_action(action_type),
            "status": status,
            "priority": _extract_numbered_field(part, "Priority"),
            "due_window": _extract_numbered_field(part, "Due Window"),
            "steps": steps,
            "dependencies": dependencies,
            "verification_checklist": verification,
            "completion": completion,
            "completion_override": override,
            "approval_required": status != "verified" and (
                completion.get("owner_signoff_needed", "").upper() == "YES" or bool(dependencies)
            ),
            "live_adapter": "wordpress_browser" if platform == "website_cms" else None,
        })
    return actions


def _parse_gbp_post_blocks(text: str) -> list[dict[str, str]]:
    blocks = re.split(r"^\s*---\s*$", text, flags=re.MULTILINE)
    posts: list[dict[str, str]] = []
    for block in blocks:
        cleaned = block.replace("**", "")
        if "DAY:" not in cleaned or "HEADLINE:" not in cleaned:
            continue
        post: dict[str, str] = {}
        for label in ("DAY", "DATE", "SERVICE", "TOPIC", "TREND_TIE", "HEADLINE", "BODY", "CAPTION", "PHOTO_FILE", "CTA", "STATUS"):
            match = re.search(rf"^{label}:\s*(.+)", cleaned, flags=re.MULTILINE)
            if match:
                post[label.lower()] = match.group(1).strip()
        if post:
            posts.append(post)
    return posts


def parse_gbp_post_actions() -> list[dict[str, Any]]:
    posts = _parse_gbp_post_blocks(_read_text(OUTPUT_DIR / "gbp_posting_schedule.md"))
    actions: list[dict[str, Any]] = []
    for index, post in enumerate(posts, start=1):
        post_id = post.get("date") or f"day-{index}"
        actions.append({
            "id": f"gbp-post-{post_id}",
            "source": "gbp_posting_schedule",
            "source_task_id": f"GBP-{index:03d}",
            "title": post.get("headline") or f"GBP post day {index}",
            "assigned_agent": "Grizzly GBP Poster Agent",
            "action_type": "publish_gbp_post",
            "platform": "google_business_profile",
            "risk": "medium",
            "status": "needs_approval" if "approval" in post.get("status", "").lower() else "dry_run_ready",
            "priority": "P2",
            "due_window": post.get("date") or "",
            "steps": [
                "Review post copy and photo selection.",
                "Publish to Google Business Profile after approval.",
            ],
            "dependencies": ["Owner approval", "Google Business Profile access"],
            "verification_checklist": [
                "Confirm post is visible on Google Business Profile.",
                "Confirm selected photo was used.",
                "Archive or mark photo as used after publishing.",
            ],
            "post": post,
            "approval_required": True,
            "live_adapter": "google_business_profile",
        })
    return actions


def _parse_facebook_post_blocks(text: str) -> list[dict[str, str]]:
    blocks = re.split(r"^\s*---\s*$", text, flags=re.MULTILINE)
    posts: list[dict[str, str]] = []
    for block in blocks:
        cleaned = block.replace("**", "")
        if "DAY:" not in cleaned or "HOOK:" not in cleaned:
            continue
        post: dict[str, str] = {}
        for label in ("DAY", "DATE", "TYPE", "SERVICE", "HOOK", "BODY", "CTA", "HASHTAGS", "PHOTO_FILE", "VIDEO_PROMPT", "STATUS"):
            match = re.search(rf"^{label}:\s*(.+)", cleaned, flags=re.MULTILINE)
            if match:
                post[label.lower()] = match.group(1).strip()
        if post:
            posts.append(post)
    return posts


def parse_facebook_post_actions() -> list[dict[str, Any]]:
    posts = _parse_facebook_post_blocks(_read_text(OUTPUT_DIR / "facebook_posting_schedule.md"))
    actions: list[dict[str, Any]] = []
    for index, post in enumerate(posts, start=1):
        post_id = post.get("date") or f"day-{index}"
        actions.append({
            "id": f"fb-post-{post_id}",
            "source": "facebook_posting_schedule",
            "source_task_id": f"FB-{index:03d}",
            "title": post.get("hook") or post.get("body", "")[:60] or f"Facebook post day {index}",
            "assigned_agent": "Grizzly Facebook Poster Agent",
            "action_type": "publish_facebook_post",
            "platform": "facebook_page",
            "risk": "medium",
            "status": "needs_approval" if "approval" in post.get("status", "").lower() else "dry_run_ready",
            "priority": "P2",
            "due_window": post.get("date") or "",
            "steps": [
                "Review post hook, body, and hashtags.",
                "For video posts: verify Gemini video prompt before publishing.",
                "Publish to Facebook Page after approval.",
            ],
            "dependencies": ["Owner approval", "Facebook Page Access Token"],
            "verification_checklist": [
                "Confirm post is visible on Facebook Business Page.",
                "For video posts: confirm video rendered correctly.",
            ],
            "post": post,
            "approval_required": True,
            "live_adapter": "facebook_page",
        })
    return actions


def _apply_approvals(actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    approvals = _load_json(ACTION_APPROVALS_FILE, {})
    for action in actions:
        approval = approvals.get(action["id"])
        if not approval:
            continue
        action["approval"] = approval
        if action["status"] == "needs_approval":
            action["status"] = "approved"
    return actions


def _load_latest_runs() -> dict[str, dict[str, Any]]:
    """Latest run record per action id, so executed/failed state survives queue rebuilds."""
    latest: dict[str, dict[str, Any]] = {}
    if not ACTION_RUN_DIR.exists():
        return latest
    for run_file in sorted(ACTION_RUN_DIR.glob("run-*.json")):
        record = _load_json(run_file, None)
        if not isinstance(record, dict) or not record.get("action_id"):
            continue
        latest[record["action_id"]] = record
    return latest


_RUN_STATUS_TO_ACTION_STATUS = {
    "live_complete": "executed",
    "live_unverified": "needs_verification",
    "adapter_failed": "failed",
}


def _apply_run_results(actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    runs = _load_latest_runs()
    for action in actions:
        record = runs.get(action["id"])
        if not record:
            continue
        action["last_run"] = {
            "id": record.get("id"),
            "live": record.get("live"),
            "status": record.get("status"),
            "message": record.get("message"),
            "created_at": record.get("created_at"),
        }
        # Owner verification still outranks everything; otherwise reflect reality.
        if action["status"] != "verified":
            mapped = _RUN_STATUS_TO_ACTION_STATUS.get(record.get("status", ""))
            if mapped:
                action["status"] = mapped
    return actions


def build_action_queue() -> dict[str, Any]:
    actions = _apply_run_results(_apply_approvals([*parse_execution_actions(), *parse_gbp_post_actions(), *parse_facebook_post_actions()]))
    adapters = {
        "wordpress_browser": wordpress_adapter_status(),
        "google_business_profile": gbp_adapter_status(),
        "facebook_page": facebook_adapter_status(),
    }
    summary = {
        "total": len(actions),
        "needs_approval": sum(1 for action in actions if action["status"] == "needs_approval"),
        "approved": sum(1 for action in actions if action["status"] == "approved"),
        "verified": sum(1 for action in actions if action["status"] == "verified"),
        "blocked_access": sum(1 for action in actions if action["status"] == "blocked_access"),
        "dry_run_ready": sum(1 for action in actions if action["status"] == "dry_run_ready"),
        "executed": sum(1 for action in actions if action["status"] == "executed"),
        "failed": sum(1 for action in actions if action["status"] == "failed"),
        "needs_verification": sum(1 for action in actions if action["status"] == "needs_verification"),
        "high_risk": sum(1 for action in actions if action["risk"] == "high"),
    }
    return {
        "version": "1.0.0",
        "generated_at": _now_iso(),
        "workflow_id": "grizzly-seo",
        "adapters": adapters,
        "summary": summary,
        "actions": actions,
    }


def write_action_queue() -> dict[str, Any]:
    payload = build_action_queue()
    _write_json(ACTION_QUEUE_FILE, payload)
    return payload


def approve_action(action_id: str, approved_by: str = "owner", note: str = "") -> dict[str, Any]:
    queue = write_action_queue()
    action = next((item for item in queue["actions"] if item["id"] == action_id), None)
    if not action:
        raise ValueError(f"Unknown action id: {action_id}")
    approvals = _load_json(ACTION_APPROVALS_FILE, {})
    approvals[action_id] = {
        "approved_by": approved_by,
        "approved_at": _now_iso(),
        "note": note,
    }
    _write_json(ACTION_APPROVALS_FILE, approvals)
    if action.get("live_adapter") == "google_business_profile":
        sync_gbp_schedule_to_workbook(dry_run=False)
        _mark_gbp_workbook_status(action, "Approved")
    return write_action_queue()


def run_action(action_id: str, live: bool = False) -> dict[str, Any]:
    queue = write_action_queue()
    action = next((item for item in queue["actions"] if item["id"] == action_id), None)
    if not action:
        raise ValueError(f"Unknown action id: {action_id}")
    if live and action.get("approval_required") and not action.get("approval"):
        result_status = "blocked_approval"
        message = "Live execution requires approval first."
        command_result = None
    elif action.get("action_type") == "gbp_profile_update":
        command_result = _run_gbp_profile_adapter(action, live=live)
        driver_result = _last_json_line(command_result.get("stdout", ""))
        command_result["driver_result"] = driver_result or None
        if command_result["exit_code"] == 0 and not live:
            result_status = "dry_run_complete"
            message = "GBP profile update dry run completed."
        elif command_result["exit_code"] == 0:
            result_status = "live_complete"
            message = f"GBP profile updated: {', '.join(driver_result.get('updates_requested', []))}"
        else:
            result_status = "adapter_failed"
            message = "GBP profile adapter failed."
    elif action.get("live_adapter") == "google_business_profile":
        command_result = _run_gbp_poster(action, live=live)
        driver_result = _last_json_line(command_result.get("stdout", ""))
        command_result["driver_result"] = driver_result or None
        if command_result["exit_code"] == 0 and not live:
            result_status = "dry_run_complete"
            message = "GBP poster dry run completed."
        elif command_result["exit_code"] == 0 and driver_result.get("verified"):
            result_status = "live_complete"
            message = "GBP post published and verified in the Posts list."
            try:
                _mark_gbp_workbook_posted(action, post_url=driver_result.get("postUrl"))
            except Exception as error:
                # The post IS live; never lose that fact over a workbook write issue.
                message = f"GBP post published and verified, but the workbook could not be updated: {error}"
        elif command_result["exit_code"] == 3 or (command_result["exit_code"] == 0 and driver_result.get("result") == "posted"):
            # Submitted but unverified: do NOT mark posted and do NOT auto-retry —
            # a blind retry can publish a duplicate.
            result_status = "live_unverified"
            message = "GBP post was submitted but could not be verified. Check the profile manually before retrying."
        else:
            result_status = "adapter_failed"
            message = "GBP poster adapter failed."
    elif action.get("live_adapter") == "facebook_page":
        command_result = _run_facebook_poster(action, live=live)
        driver_result = _last_json_line(command_result.get("stdout", ""))
        command_result["driver_result"] = driver_result or None
        if command_result["exit_code"] == 0 and not live:
            result_status = "dry_run_complete"
            message = "Facebook poster dry run completed."
        elif command_result["exit_code"] == 0 and driver_result.get("status") == "success":
            result_status = "live_complete"
            message = f"Facebook post published. Post ID: {driver_result.get('post_id')}"
        else:
            result_status = "adapter_failed"
            message = f"Facebook poster adapter failed: {driver_result.get('message', command_result.get('stderr', ''))}"
    elif action.get("live_adapter") == "wordpress_browser":
        command_result = _run_wordpress_adapter(action, live=live)
        result_status = "live_complete" if live and command_result["exit_code"] == 0 else "dry_run_complete" if command_result["exit_code"] == 0 else "adapter_failed"
        message = "WordPress adapter completed." if command_result["exit_code"] == 0 else "WordPress adapter failed."
    elif live:
        result_status = "blocked_adapter"
        message = f"No live adapter configured for {action['platform']} yet."
        command_result = None
    else:
        result_status = "dry_run_complete"
        message = "Dry run generated execution payload only. No live system was changed."
        command_result = None

    run_record = {
        "id": f"run-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}-{action_id}",
        "action_id": action_id,
        "live": live,
        "status": result_status,
        "message": message,
        "action": action,
        "command_result": command_result,
        "created_at": _now_iso(),
    }
    ACTION_RUN_DIR.mkdir(exist_ok=True)
    _write_json(ACTION_RUN_DIR / f"{run_record['id']}.json", run_record)
    return run_record


def _load_wordpress_config() -> dict[str, Any]:
    if not WORDPRESS_SITE_CONFIG.exists():
        raise FileNotFoundError(f"WordPress site config not found: {WORDPRESS_SITE_CONFIG}")
    return json.loads(WORDPRESS_SITE_CONFIG.read_text(encoding="utf-8"))


def wordpress_adapter_status() -> dict[str, Any]:
    status: dict[str, Any] = {
        "name": "wordpress-browser",
        "site_config": str(WORDPRESS_SITE_CONFIG),
        "adapter": WORDPRESS_ACTION_ADAPTER or None,
        "browser_session_dir": str(WORDPRESS_BROWSER_SESSION_DIR),
        "state": "missing",
        "config_ready": False,
        "browser_session_ready": WORDPRESS_BROWSER_SESSION_DIR.exists(),
        "live_adapter_ready": False,
        "capabilities": [],
        "missing": [],
    }
    if not WORDPRESS_SITE_CONFIG.exists():
        status["missing"].append("WordPress site config")
        return status
    try:
        config = _load_wordpress_config()
    except Exception as error:
        status["state"] = "error"
        status["missing"].append(str(error))
        return status
    status["config_ready"] = True
    status["site_id"] = config.get("site_id")
    status["site_url"] = config.get("site_url")
    status["wp_admin_url"] = config.get("wp_admin_url")
    status["contact_forms"] = [
        {"name": form.get("name"), "post_id": form.get("post_id")}
        for form in config.get("contact_forms", [])
    ]
    status["capabilities"] = [
        "wp_session_check",
        "cf7_inventory",
        "cf7_mail_settings_dry_run",
        "cf7_mail_edit",
        "public_form_submit_test",
        "wordpress_page_update_draft",
    ]
    if WORDPRESS_ACTION_ADAPTER:
        adapter_path = Path(WORDPRESS_ACTION_ADAPTER)
        status["live_adapter_ready"] = adapter_path.exists()
        if not adapter_path.exists():
            status["missing"].append("WordPress action adapter script")
    else:
        status["missing"].append("WORDPRESS_ACTION_ADAPTER not configured")
    if not status["browser_session_ready"]:
        status["missing"].append("WordPress browser session directory")
    if status["config_ready"] and status["live_adapter_ready"] and status["browser_session_ready"]:
        status["state"] = "live_ready"
    elif status["config_ready"]:
        status["state"] = "approval_ready"
    return status


def _run_wordpress_adapter(action: dict[str, Any], live: bool) -> dict[str, Any]:
    try:
        config = _load_wordpress_config()
    except Exception as error:
        return {
            "exit_code": 2,
            "command": "wordpress-browser",
            "stdout": "",
            "stderr": str(error),
        }
    payload = {
        "site": {
            "site_id": config.get("site_id"),
            "site_url": config.get("site_url"),
            "wp_admin_url": config.get("wp_admin_url"),
        },
        "live": live,
        "action": action,
    }
    if not WORDPRESS_ACTION_ADAPTER:
        return {
            "exit_code": 127,
            "command": "wordpress-browser",
            "stdout": "",
            "stderr": "WORDPRESS_ACTION_ADAPTER is not configured yet.",
        }
    adapter_path = Path(WORDPRESS_ACTION_ADAPTER)
    if not adapter_path.exists():
        return {
            "exit_code": 127,
            "command": str(adapter_path),
            "stdout": "",
            "stderr": f"WordPress action adapter not found: {adapter_path}",
        }
    command = [
        "node",
        str(adapter_path),
        "--config",
        str(WORDPRESS_SITE_CONFIG),
        "--payload",
        json.dumps(payload),
    ]
    return _run_subprocess(
        command,
        cwd=str(adapter_path.parent),
        timeout_s=WORDPRESS_ADAPTER_TIMEOUT_S,
        display_command=" ".join(command[:4]) + " --payload <json>",
    )


def _run_gbp_poster(action: dict[str, Any], live: bool) -> dict[str, Any]:
    if not GBP_POSTER_SCRIPT.exists():
        return {
            "exit_code": 127,
            "command": str(GBP_POSTER_SCRIPT),
            "stdout": "",
            "stderr": f"GBP poster script not found: {GBP_POSTER_SCRIPT}",
        }
    date_value = action.get("post", {}).get("date") or action.get("due_window")
    if not date_value:
        return {
            "exit_code": 2,
            "command": str(GBP_POSTER_SCRIPT),
            "stdout": "",
            "stderr": "GBP post action is missing a date.",
        }
    command = ["node", str(GBP_POSTER_SCRIPT), "--date", date_value, "--config", str(GBP_POSTER_CONFIG)]
    if not live:
        command.append("--dry-run")
    if live and GBP_POSTER_HEADLESS:
        command.append("--headless")
    return _run_subprocess(
        command,
        cwd=str(GBP_POSTER_SCRIPT.parent),
        timeout_s=GBP_POSTER_TIMEOUT_S,
    )


def _run_gbp_profile_adapter(action: dict[str, Any], live: bool) -> dict[str, Any]:
    adapter_path = Path(GBP_PROFILE_ADAPTER)
    if not adapter_path.exists():
        return {
            "exit_code": 127,
            "command": str(adapter_path),
            "stdout": "",
            "stderr": f"GBP profile adapter not found: {adapter_path}",
        }
    payload = {"live": live, "action": action}
    command = ["node", str(adapter_path), "--payload", json.dumps(payload)]
    if not live:
        command.append("--dry-run")
    return _run_subprocess(
        command,
        cwd=str(adapter_path.parent),
        timeout_s=GBP_PROFILE_ADAPTER_TIMEOUT_S,
        display_command=" ".join(command[:3]) + " --payload <json>",
    )


def _run_facebook_poster(action: dict[str, Any], live: bool) -> dict[str, Any]:
    adapter_path = Path(FACEBOOK_POSTER_ADAPTER)
    if not adapter_path.exists():
        return {
            "exit_code": 127,
            "command": str(adapter_path),
            "stdout": "",
            "stderr": f"Facebook poster adapter not found: {adapter_path}",
        }
    payload = {"live": live, "action": action}
    command = ["node", str(adapter_path), "--payload", json.dumps(payload)]
    if not live:
        command.append("--dry-run")
    return _run_subprocess(
        command,
        cwd=str(adapter_path.parent),
        timeout_s=FACEBOOK_POSTER_TIMEOUT_S,
        display_command=" ".join(command[:3]) + " --payload <json>",
    )


def facebook_adapter_status() -> dict[str, Any]:
    page_id = os.getenv("FB_PAGE_ID", "").strip()
    access_token = os.getenv("FB_PAGE_ACCESS_TOKEN", "").strip()
    adapter_path = Path(FACEBOOK_POSTER_ADAPTER)
    missing = []
    if not page_id:
        missing.append("FB_PAGE_ID not set in .env")
    if not access_token:
        missing.append("FB_PAGE_ACCESS_TOKEN not set in .env")
    if not adapter_path.exists():
        missing.append("Facebook Graph API adapter script not found")
    state = "live_ready" if (page_id and access_token and adapter_path.exists()) else "blocked"
    return {
        "name": "facebook-graph-api",
        "adapter": str(adapter_path),
        "page_id_set": bool(page_id),
        "access_token_set": bool(access_token),
        "state": state,
        "missing": missing,
    }


def _load_gbp_config() -> dict[str, Any]:
    if not GBP_POSTER_CONFIG.exists():
        raise FileNotFoundError(f"GBP poster config not found: {GBP_POSTER_CONFIG}")
    return json.loads(GBP_POSTER_CONFIG.read_text(encoding="utf-8"))


def gbp_adapter_status() -> dict[str, Any]:
    status: dict[str, Any] = {
        "name": "gbp-browser-poster",
        "script": str(GBP_POSTER_SCRIPT),
        "config": str(GBP_POSTER_CONFIG),
        "state": "missing",
        "browser_session_ready": False,
        "media_upload_ready": False,
        "workbook_ready": False,
        "photo_folder_ready": False,
        "missing": [],
    }
    if not GBP_POSTER_SCRIPT.exists():
        status["missing"].append("GBP poster script")
        return status
    if not GBP_POSTER_CONFIG.exists():
        status["missing"].append("GBP poster config")
        return status
    try:
        config = _load_gbp_config()
    except Exception as error:
        status["state"] = "error"
        status["missing"].append(str(error))
        return status

    status["browser_session_ready"] = GBP_BROWSER_SESSION_DIR.exists()
    status["media_upload_ready"] = all(
        bool(config.get(key))
        for key in ("supabase_url", "supabase_service_role_key", "supabase_bucket")
    )
    status["workbook_ready"] = bool(config.get("workbook_path")) and Path(config["workbook_path"]).exists()
    status["photo_folder_ready"] = bool(config.get("curated_photo_folder")) and Path(config["curated_photo_folder"]).exists()

    if not status["browser_session_ready"]:
        status["missing"].append("Google Business Profile browser session")
    if not status["media_upload_ready"]:
        status["missing"].append("Supabase media upload config")
    if not status["workbook_ready"]:
        status["missing"].append("GBP approval workbook")
    if not status["photo_folder_ready"]:
        status["missing"].append("Curated GBP photo folder")

    if status["browser_session_ready"] and status["workbook_ready"] and status["photo_folder_ready"]:
        status["state"] = "live_ready"
    elif status["workbook_ready"] and status["photo_folder_ready"]:
        status["state"] = "approval_ready"
    else:
        status["state"] = "blocked"
    return status


def _open_gbp_workbook():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to sync the GBP workbook.") from exc
    config = _load_gbp_config()
    workbook_path = Path(config["workbook_path"])
    if not workbook_path.exists():
        raise FileNotFoundError(f"GBP workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path)
    if "Posts" not in workbook.sheetnames:
        raise RuntimeError("GBP workbook must contain a Posts sheet.")
    sheet = workbook["Posts"]
    headers = [cell.value for cell in sheet[1]]
    missing = [header for header in GBP_WORKBOOK_HEADERS if header not in headers]
    if missing:
        raise RuntimeError(f"GBP workbook missing headers: {missing}")
    columns = {header: headers.index(header) + 1 for header in GBP_WORKBOOK_HEADERS}
    return config, workbook_path, workbook, sheet, columns


def _save_gbp_workbook(workbook: Any, workbook_path: Path) -> None:
    try:
        workbook.save(workbook_path)
    except PermissionError as exc:
        raise RuntimeError(
            f"Could not save GBP workbook — it is likely open in Excel. Close it and retry: {workbook_path}"
        ) from exc


def _row_date(value: Any) -> str:
    if hasattr(value, "date"):
        return value.date().isoformat()
    if isinstance(value, str):
        return value[:10]
    return ""


def _find_workbook_row(sheet: Any, columns: dict[str, int], date_value: str) -> int | None:
    for row in range(2, sheet.max_row + 1):
        if _row_date(sheet.cell(row, columns["Date"]).value) == date_value:
            return row
    return None


def _caption_for_post(post: dict[str, str]) -> str:
    def clean_gbp_text(value: str) -> str:
        # Google Business Profile rejects post descriptions that include phone numbers.
        value = re.sub(r"\+?1?[\s.-]*\(?\d{3}\)?[\s.-]*\d{3}[\s.-]*\d{4}", "", value)
        value = re.sub(r"\s{2,}", " ", value)
        value = re.sub(r"\s+([.!?,])", r"\1", value)
        return value.strip()

    parts = [
        clean_gbp_text(post.get("headline", "")),
        clean_gbp_text(post.get("body", "")),
        clean_gbp_text(post.get("cta", "")),
    ]
    return "\n\n".join(part for part in parts if part)


def sync_gbp_schedule_to_workbook(dry_run: bool = False) -> dict[str, Any]:
    config, workbook_path, workbook, sheet, columns = _open_gbp_workbook()
    curated_folder = Path(config.get("curated_photo_folder", ""))
    posts = _parse_gbp_post_blocks(_read_text(OUTPUT_DIR / "gbp_posting_schedule.md"))
    updates: list[dict[str, Any]] = []
    next_new_row = sheet.max_row + 1
    for post in posts:
        date_value = post.get("date", "")
        if not date_value:
            continue
        row = _find_workbook_row(sheet, columns, date_value)
        is_new = row is None
        if row is None:
            row = next_new_row
            next_new_row += 1
        photo_file = post.get("photo_file", "")
        photo_path = curated_folder / photo_file if photo_file else Path("")
        updates.append({"date": date_value, "row": row, "new": is_new, "title": post.get("headline", "")})
        if dry_run:
            continue
        existing_status = str(sheet.cell(row, columns["Status"]).value or "").strip()
        existing_posted = bool(sheet.cell(row, columns["Posted"]).value)
        existing_image_link = sheet.cell(row, columns["ImageLink"]).value
        existing_post_url = sheet.cell(row, columns["GBPPostUrl"]).value
        existing_posted_at = sheet.cell(row, columns["PostedAt"]).value
        sheet.cell(row, columns["Date"]).value = date_value
        sheet.cell(row, columns["PostType"]).value = "STANDARD"
        sheet.cell(row, columns["Topic"]).value = post.get("topic") or post.get("service") or post.get("headline")
        sheet.cell(row, columns["AssetSource"]).value = "Workspace Shared"
        sheet.cell(row, columns["AssetIdOrDescription"]).value = str(photo_path) if photo_file else ""
        sheet.cell(row, columns["CTA"]).value = post.get("cta", "")
        sheet.cell(row, columns["Status"]).value = existing_status if existing_status in {"Approved", "Posted"} else post.get("status") or "Needs approval"
        sheet.cell(row, columns["CaptionDraft"]).value = _caption_for_post(post)
        sheet.cell(row, columns["ImageLink"]).value = existing_image_link
        sheet.cell(row, columns["Posted"]).value = existing_posted
        sheet.cell(row, columns["PostedAt"]).value = existing_posted_at
        sheet.cell(row, columns["GBPPostUrl"]).value = existing_post_url
        sheet.cell(row, columns["Notes"]).value = f"Synced from SEO Agents action queue at {_now_iso()}; {post.get('trend_tie', '')}"
    backup_path = None
    if not dry_run:
        backup_path = workbook_path.with_suffix(f".backup-seo-sync-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup_path)
        _save_gbp_workbook(workbook, workbook_path)
    return {
        "workbook_path": str(workbook_path),
        "backup_path": str(backup_path) if backup_path else None,
        "dry_run": dry_run,
        "posts_found": len(posts),
        "updates": updates,
    }


def mark_gbp_dates_approved(dates: list[str]) -> dict[str, Any]:
    """Stamp the workbook Status to 'Approved' for the given post dates.

    Weekly approval is a single decision: when the week is approved, every day in
    that week should post on its scheduled date. The GBP poster gates on the
    workbook Status column, so this propagates the one weekly approval to all
    days' rows. Rows already 'Posted' are left untouched.
    """
    _, workbook_path, workbook, sheet, columns = _open_gbp_workbook()
    approved: list[str] = []
    skipped: list[str] = []
    for date_value in dates:
        if not date_value:
            continue
        row = _find_workbook_row(sheet, columns, date_value)
        if not row:
            skipped.append(date_value)
            continue
        existing = str(sheet.cell(row, columns["Status"]).value or "").strip()
        if existing == "Posted":
            skipped.append(date_value)
            continue
        sheet.cell(row, columns["Status"]).value = "Approved"
        sheet.cell(row, columns["Notes"]).value = f"Approved (weekly) at {_now_iso()}"
        approved.append(date_value)
    if approved:
        _save_gbp_workbook(workbook, workbook_path)
    return {"workbook_path": str(workbook_path), "approved": approved, "skipped": skipped}


def _mark_gbp_workbook_status(action: dict[str, Any], status: str) -> None:
    date_value = action.get("post", {}).get("date") or action.get("due_window")
    if not date_value:
        return
    _, workbook_path, workbook, sheet, columns = _open_gbp_workbook()
    row = _find_workbook_row(sheet, columns, date_value)
    if not row:
        return
    sheet.cell(row, columns["Status"]).value = status
    sheet.cell(row, columns["Notes"]).value = f"{status} from SEO Agents action queue at {_now_iso()}"
    _save_gbp_workbook(workbook, workbook_path)


def _mark_gbp_workbook_posted(action: dict[str, Any], post_url: str | None = None) -> None:
    date_value = action.get("post", {}).get("date") or action.get("due_window")
    if not date_value:
        return
    _, workbook_path, workbook, sheet, columns = _open_gbp_workbook()
    row = _find_workbook_row(sheet, columns, date_value)
    if not row:
        return
    sheet.cell(row, columns["Status"]).value = "Posted"
    sheet.cell(row, columns["Posted"]).value = True
    sheet.cell(row, columns["PostedAt"]).value = _now_iso()
    if post_url:
        sheet.cell(row, columns["GBPPostUrl"]).value = post_url
    sheet.cell(row, columns["Notes"]).value = "Posted and verified by MCC browser adapter."
    _save_gbp_workbook(workbook, workbook_path)


def format_action_queue_text(queue: dict[str, Any]) -> str:
    summary = queue["summary"]
    lines = [
        f"Actions: {summary['total']}",
        f"Needs approval: {summary['needs_approval']}",
        f"Approved: {summary['approved']}",
        f"Blocked access: {summary['blocked_access']}",
        f"Dry-run ready: {summary['dry_run_ready']}",
        "",
    ]
    for action in queue["actions"]:
        lines.append(
            f"- {action['id']} [{action['status']}] {action['title']} "
            f"({action['assigned_agent']} / {action['platform']} / {action['risk']})"
        )
    return "\n".join(lines)
